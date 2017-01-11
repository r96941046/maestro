import datetime
import tkinter as tk
from tkinter.filedialog import askopenfilename
from collections import defaultdict

from openpyxl import load_workbook


class Maestro():

    COL_PRICE = 18
    COL_INVOICE = 26

    def __init__(self):

        self.window = tk.Tk()

        self.window.title('Maestro')

        tk.Label(self.window, text="Excel file").grid(row=0, column=0)
        tk.Label(self.window, text="Column number of the price").grid(row=1, column=0)
        tk.Label(self.window, text="Column number of the invoice").grid(row=2, column=0)
        tk.Label(self.window, text="Target price").grid(row=3, column=0)

        self.s_filename = tk.StringVar()
        self.e_filemame = tk.Entry(self.window, text=self.s_filename)
        self.e_filemame.grid(row=0, column=1)
        self.b_filename = tk.Button(self.window, text='Browse', command=self.browse)
        self.b_filename.grid(row=0, column=2)

        self.i_col_price = tk.IntVar()
        self.i_col_price.set(self.COL_PRICE)

        self.i_col_invoice = tk.IntVar()
        self.i_col_invoice.set(self.COL_INVOICE)

        self.e_col_price = tk.Entry(self.window, text=self.i_col_price)
        self.e_col_price.grid(row=1, column=1)

        self.e_col_invoice = tk.Entry(self.window, text=self.i_col_invoice)
        self.e_col_invoice.grid(row=2, column=1)

        self.e_target = tk.Entry(self.window)
        self.e_target.grid(row=3, column=1)

        self.window.grid_rowconfigure(4, minsize=15)

        tk.Label(self.window, text="Total valid rows").grid(row=5, column=0)
        self.l_valid_rows = tk.Label(self.window, text="")
        self.l_valid_rows.grid(row=5, column=1)

        tk.Label(self.window, text="Calculate time").grid(row=6, column=0)
        self.l_calculate_time = tk.Label(self.window, text="")
        self.l_calculate_time.grid(row=6, column=1)

        tk.Label(self.window, text="Price subsets").grid(row=7, column=0)
        self.s_price_subsets = tk.StringVar()
        self.l_price_subsets = tk.Entry(self.window, text=self.s_price_subsets)
        self.l_price_subsets.grid(row=7, column=1)

        tk.Label(self.window, text="Invoice subsets").grid(row=8, column=0)
        self.s_invoice_subsets = tk.StringVar()
        self.l_invoice_subsets = tk.Entry(self.window, text=self.s_invoice_subsets)
        self.l_invoice_subsets.grid(row=8, column=1)

        self.window.grid_rowconfigure(9, minsize=15)

        self.s_calculate = tk.StringVar()
        self.s_calculate.set('Calculate')
        self.b_calculate = tk.Button(self.window, text=self.s_calculate.get(), command=self.calculate)
        self.b_calculate.grid(row=10, column=0)

        self.b_cancel = tk.Button(self.window, text='Cancel', command=self.cancel)
        self.b_cancel.grid(row=10, column=1)

        self.window.mainloop()

    def browse(self):

        self.s_filename.set(askopenfilename())

    def cancel(self):

        self.window.destroy()

    def subsets_with_sum(self, lst, target):

        x = 1

        def _a(idx, l, r, t):

            if t == sum(l): r.append(l)
            elif t < sum(l): return
            for u in range(idx, len(lst)):
                _a(u + x, l + [lst[u]], r, t)
            return r

        return _a(0, [], [], target)

    def calculate(self):

        start = datetime.datetime.now()

        references = defaultdict(list)
        prices = []

        input_price_sum = self.e_target.get()

        try:
            price_sum = float(input_price_sum)
        except ValueError:
            print('Target price must be a number.')

        with open(self.s_filename.get(), 'rb') as f:

            wb = load_workbook(f)
            ws = wb.get_active_sheet()

            for row in ws.iter_rows(row_offset=1):

                price = row[self.i_col_price.get()].value

                if price is not None and price > 0 and price <= price_sum:

                    invoice = row[self.i_col_invoice.get()].value
                    prices.append(price)
                    references[price].append(invoice)

            prices.sort()
            self.l_valid_rows['text'] = str(len(prices))

            price_subsets = set([tuple(s) for s in self.subsets_with_sum(prices, price_sum)])
            invoice_subsets = [tuple(references[p] for p in ps) for ps in price_subsets]

            s_price_subset = str(price_subsets) if len(price_subsets) else '()'
            self.s_price_subsets.set(s_price_subset)
            self.s_invoice_subsets.set(str(invoice_subsets))

            end = datetime.datetime.now()
            self.l_calculate_time['text'] = str(end - start)


if __name__ == '__main__':

    maestro = Maestro()
